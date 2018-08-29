#-*- coding:utf-8 -*-


from __future__ import print_function
import sys
import re
import pyodbc
import openpyxl


def main():
    DBfile = r"./WSN_20180828_221805.accdb"  # 数据库文件
    conn = pyodbc.connect(r"Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=" + DBfile + ";Uid=;Pwd=;")
    cursor = conn.cursor()
    SQL = 'SELECT * from Results;'
    cursor.execute(SQL)
    Resultslist = cursor.fetchall()
    SQL = 'SELECT * from Compute;'
    cursor.execute(SQL)
    Computelist = cursor.fetchall()
    pattern = r'(.+?)\.'
    xlsname = "".join(re.findall(pattern, DBfile, flags=re.IGNORECASE)) +'.xlsx'
    savexls(Resultslist, Computelist, xlsname)


def savexls(Resultslist, Computelist, filename):
    wb = openpyxl.Workbook()

    # Compute sheet
    ws = wb.worksheets[0]
    ws.title = 'Compute'
    ws['A1'] = 'ID'
    ws['B1'] = 'nodeId'
    ws['C1'] = 'lossPkgCount'
    ws['D1'] = 'lossPkgRate'
    ws['E1'] = 'avgRSSI'
    ws['F1'] = 'voltDiff'
    ws['G1'] = 'avgHopCount'
    ws['H1'] = 'TimeDiff'
    for i, x in enumerate(Computelist):
        try:
            ws['A' + str(i + 2)] = x[0]
            # print(type(x[1]))
            ws['B' + str(i + 2)] = hex(int(x[1]))  # "".join(field[2:4])
            ws['C' + str(i + 2)] = x[2]  # float(RSSI_Val)  # gettime # RSSI need to convert
            ws['D' + str(i + 2)] = x[3]  # int("".join(field[4:6]), 16)  # seqNo
            ws['E' + str(i + 2)] = x[4]  # datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ws['F' + str(i + 2)] = x[5]  # Voltage  # "".join(field[7:9]) # voltage need to convert
            ws['G' + str(i + 2)] = x[6]  # int("".join(field[9]), 16)  # hopCount
            ws['H' + str(i + 2)] = x[7]  # timedifference  # "".join(field[10:]) # timedifference need to convert
        except IndexError as err:
            print("out of the field range of endline. " + str(err))
            break


    # Results sheet
    wb.create_sheet(index=0, title='Results')
    ws = wb['Results']
    ws['A1'] = 'id'
    ws['B1'] = 'nodeId'
    ws['C1'] = 'rssi'
    ws['D1'] = 'seqNO'
    ws['E1'] = 'gettime'
    ws['F1'] = 'volt'
    ws['G1'] = 'hopCount'
    ws['H1'] = 'timeDifference'


    for i, x in enumerate(Resultslist):
        try:
            ws['A' + str(i + 2)] = x[0]
            ws['B' + str(i + 2)] = hex(int(x[1]))  # "".join(field[2:4])
            ws['C' + str(i + 2)] = x[2]  # float(RSSI_Val)  # gettime # RSSI need to convert
            ws['D' + str(i + 2)] = x[3]  # int("".join(field[4:6]), 16)  # seqNo
            ws['E' + str(i + 2)] = x[4]  # datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ws['F' + str(i + 2)] = x[5]  # Voltage  # "".join(field[7:9]) # voltage need to convert
            ws['G' + str(i + 2)] = x[6]  # int("".join(field[9]), 16)  # hopCount
            ws['H' + str(i + 2)] = x[7]  # timedifference  # "".join(field[10:]) # timedifference need to convert
        except IndexError as err:
            print("out of the field range of endline. " + str(err))
            break


    wb.save(filename)


if __name__ == '__main__':
    main()