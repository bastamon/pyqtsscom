# -*-coding:utf-8-*-
from __future__ import print_function
from openpyxl import load_workbook
import sqlite3
import re


def convertRSSI(filename):
    wb = load_workbook(filename)
    ws1 = wb['Sheet1']
    sqlfile = "".join(re.findall(r'(.+?)\.', filename, flags=re.IGNORECASE)) + ".db"
    con = sqlite3.connect(sqlfile)
    cur = con.cursor()
    try:
        cur.execute('''create table if not exists Results
        (id             INTEGER     PRIMARY KEY     AUTOINCREMENT,
         nodeId         CHAR(2)     NOT NULL,
         rssi           INT         NOT NULL,
         seqNo          INT         NOT NULL,
         gettime        TEXT        NOT NULL,
         volt           FLOAT       NOT NULL,
         hopCount       INT         NOT NULL,
         timeDifference FLOAT       NOT NULL
         );
        ''')
    except sqlite3.Error as e:
        print('sqlite3 error occur In create: %s', e.args[0])


    for i, row in enumerate(list(ws1.rows)[1:]):
        x = int(row[1].value)
        # print(x)
        if x > 128:
            ws1['B' + str(i + 2)] = x - 45 - 256
        elif x > 0:
            ws1['B' + str(i + 2)] = x - 45
        # savesql(list(row), sqlfile)
        content = list(row)
        temp = [i.value for i in content[0:7]]
        try:
            cur.execute('''insert into Results(nodeId,rssi,seqNo,gettime,volt,hopCount,timeDifference) 
                                     values(?,?,?,?,?,?,?)''', temp)
        except sqlite3.Error as e:
            print("An sqlite3 error occurred insert: %s", e.args[0])
        # finally:
    con.commit()
    cur.close()
    con.close()


    Statistic(sqlfile)
    # Compute sheet
    wb.create_sheet(index=0, title='Compute')
    ws1 = wb['Compute']
    ws1['A1'] = 'ID'
    ws1['B1'] = 'nodeId'
    ws1['C1'] = 'lostPktCount'
    ws1['D1'] = 'lostPktRatio'
    ws1['E1'] = 'avgRSSI'
    ws1['F1'] = 'voltDiff'
    ws1['G1'] = 'avgHopCount'
    ws1['H1'] = 'avgTimeDiff'

    try:
        con = sqlite3.connect(sqlfile)
        cur = con.cursor()
        statistic_data = cur.execute("SELECT * FROM Statistic;").fetchall()
        for i, x in enumerate(statistic_data):
            ws1['A' + str(i + 2)] = x[0]
            ws1['B' + str(i + 2)] = x[1]
            ws1['C' + str(i + 2)] = x[2]
            ws1['D' + str(i + 2)] = x[3]
            ws1['E' + str(i + 2)] = x[4]
            ws1['F' + str(i + 2)] = x[5]
            ws1['G' + str(i + 2)] = x[6]
            ws1['H' + str(i + 2)] = x[7]
        cur.close()
        con.close()
    except sqlite3.Error as e:
        print('sqlite3 error occur In extract: %s', e.args[0])
    wb.save(filename)


def savesql(content, filename):
    con = sqlite3.connect(filename)
    cur = con.cursor()
    try:
        cur.execute('''create table if not exists Results
        (id             INTEGER     PRIMARY KEY     AUTOINCREMENT,
         nodeId         CHAR(2)     NOT NULL,
         rssi           INT         NOT NULL,
         seqNo          INT         NOT NULL,
         gettime        TEXT        NOT NULL,
         volt           FLOAT       NOT NULL,
         hopCount       INT         NOT NULL,
         timeDifference FLOAT       NOT NULL
         );
        ''')
    except sqlite3.Error as e:
        print('sqlite3 error occur In create: %s', e.args[0])

    # nodeId = content[0].value
    # rssi = content[1].value
    # seqNo = content[2].value
    # gettime = content[3].value
    # volt = content[4].value
    # hopCount = content[5].value
    # timedifference = content[6].value
    temp = [content[0].value, content[1].value, content[2].value, content[3].value, content[4].value, content[5].value, content[6].value]
    try:
        cur.execute('''insert into Results(nodeId,rssi,seqNo,gettime,volt,hopCount,timeDifference) 
                                values(?,?,?,?,?,?,?)''', temp)
    except sqlite3.Error as e:
        print("An sqlite3 error occurred insert: %s", e.args[0])
    finally:
        con.commit()
        cur.close()
        con.close()


def Statistic(filename):
    con = sqlite3.connect(filename)
    cur = con.cursor()
    try:
        cur.execute('''create table if not exists Statistic
        (ID             INTEGER     PRIMARY KEY     AUTOINCREMENT,
         nodeId         CHAR(2)     NOT NULL,
         lossPkgCount   INT         NOT NULL,
         lossPkgRate    FLOAT       NOT NULL,
         avgRSSI        FLOAT       NOT NULL,
         voltDiff       FLOAT       NOT NULL,
         avgHopCount    INT         NOT NULL,
         TimeDiff       FLOAT       NOT NULL
         );
        ''')
    except sqlite3.Error as e:
        print('sqlite3 error occur In create: %s', e.args[0])

    rows = cur.execute("SELECT * FROM Results ORDER BY nodeId ASC,seqNo ASC;").fetchall()
    row_seq = []
    row_seq.append(rows[0][1])
    for i in range(1, len(rows)):
        if rows[i][1] != rows[i - 1][1]:
            row_seq.append(rows[i][1])

    for i, x in enumerate(row_seq):
        x = str(x)
        max_volt = cur.execute("SELECT max(volt) FROM Results WHERE nodeId = '" + x + "';").fetchall()[0][0]
        min_volt = cur.execute("SELECT min(volt) FROM Results WHERE nodeId = '" + x + "';").fetchall()[0][0]
        voltDiff = max_volt - min_volt
        avgRSSI = cur.execute("SELECT avg(rssi) FROM Results WHERE nodeId = '" + x + "';").fetchall()[0][0]
        avgHopCount = cur.execute("SELECT avg(hopCount) FROM Results WHERE nodeId = '" + x + "';").fetchall()[0][0]
        TimeDiff = cur.execute("SELECT avg(timeDifference) FROM Results WHERE nodeId = '" + x + "';").fetchall()[0][0]
        recvCount = cur.execute("SELECT count(*) FROM Results WHERE nodeId = '" + x + "';").fetchall()[0][0]
        lossPkgCount = 0
        seq_data = cur.execute("SELECT seqNo FROM Results WHERE nodeId = '" + x + "' ORDER BY seqNo ASC;").fetchall()
        for j in range(1, len(seq_data)):
            # < 排除了可能重发而导致的逻辑丢包计数
            if seq_data[j - 1][0] + 1 < seq_data[j][0]:
                lossPkgCount += seq_data[j][0] - seq_data[j - 1][0] - 1

        row = [0 for _ in range(7)]
        row[0] = x  # nodeId
        row[1] = lossPkgCount  # lossPkgCount
        row[2] = 100.0 * float(lossPkgCount) / (recvCount + lossPkgCount)  # lossPkgRate
        row[3] = avgRSSI
        row[4] = voltDiff
        row[5] = avgHopCount
        row[6] = TimeDiff
        cur.execute(
            '''insert into Statistic(nodeId,lossPkgCount,lossPkgRate,avgRSSI,voltDiff,avgHopCount,TimeDiff) values(?,?,?,?,?,?,?)''',
            tuple(row))
    con.commit()
    cur.close()
    con.close()
    print("statistic complete")


def main():
    convertRSSI(u'工作簿1.xlsx')


if __name__ == '__main__':
    main()
