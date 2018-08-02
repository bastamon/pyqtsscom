# -*-coding:utf-8-*-
from __future__ import print_function
import sys
import re
import openpyxl
import datetime
import sqlite3


def read(filename, protId='10'):
    with open(filename) as readfile:
        content = readfile.read()
    content1 = content.split(" ")
    content2 = ""
    for i in range(len(content1)):
        # content1[i] is frameHead, content1[i + 1] is protId recognized code
        if content1[i] == '44' and content1[i + 1] == protId:
            content2 += " ".join(content1[i:i + 13])  # print("\t".join(content1[i:]))
            content2 += '\n'
    assert content2 != '', "file can't be empty"
    return content2


def analyze(filename, protId='10'):
    strvar = ""
    strvar += read(filename, protId)

    pattern = r'(.+?)\.'
    pathname = "".join(re.findall(pattern, filename, flags=re.IGNORECASE))
    xlspath = pathname + ".xlsx"
    xlsfile = xlspath

    sqlpath = pathname + ".db"
    sqlfile = sqlpath

    print(sqlfile)
    print(xlsfile)

    content = strvar.split('\n')
    savesql(content, sqlfile)
    Statistic(sqlfile)
    savexls(content, xlsfile)


def savexls(content, filename):
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    # Results sheet
    ws.title = 'Results'
    ws['A1'] = 'id'
    ws['B1'] = 'nodeId'
    ws['C1'] = 'rssi'
    ws['D1'] = 'seqNO'
    ws['E1'] = 'gettime'
    ws['F1'] = 'volt'
    ws['G1'] = 'hopCount'
    ws['H1'] = 'timeDifference'

    # 计算从数据库查询得
    pattern = r'(.+?)\.'
    pathname = "".join(re.findall(pattern, filename, flags=re.IGNORECASE))
    sqlfile = pathname + ".db"

    con = sqlite3.connect(sqlfile)
    cur = con.cursor()
    results_dara = cur.execute("SELECT * FROM Results;").fetchall()

    for i, x in enumerate(results_dara):
        try:
            ws['A' + str(i + 2)] = x[0]
            ws['B' + str(i + 2)] = x[1]  # "".join(field[2:4])
            ws['C' + str(i + 2)] = x[2]  # float(RSSI_Val)  # gettime # RSSI need to convert
            ws['D' + str(i + 2)] = x[3]  # int("".join(field[4:6]), 16)  # seqNo
            ws['E' + str(i + 2)] = x[4]  # datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ws['F' + str(i + 2)] = x[5]  # Voltage  # "".join(field[7:9]) # voltage need to convert
            ws['G' + str(i + 2)] = x[6]  # int("".join(field[9]), 16)  # hopCount
            ws['H' + str(i + 2)] = x[7]  # timedifference  # "".join(field[10:]) # timedifference need to convert
        except IndexError as err:
            print("out of the field range of endline. " + str(err))
            break
    cur.close()
    con.close()

    # Results_Statistic sheet
    wb.create_sheet(index=0, title='Results_Statistic')
    ws1 = wb['Results_Statistic']
    ws1['A1'] = 'ID'
    ws1['B1'] = 'nodeId'
    ws1['C1'] = 'lossPkgCount'
    ws1['D1'] = 'lossPkgRate'
    ws1['E1'] = 'avgRSSI'
    ws1['F1'] = 'voltDiff'
    ws1['G1'] = 'avgHopCount'
    ws1['H1'] = 'TimeDiff'

    try:
        con = sqlite3.connect(sqlfile)
        cur = con.cursor()
        statistic_data = cur.execute("SELECT * FROM Statistic;").fetchall()
        for i, x in enumerate(statistic_data):
            ws1['A' + str(i + 2)] = x[0]
            ws1['B' + str(i + 2)] = x[1]
            ws1['C' + str(i + 2)] = x[2]
            ws1['D' + str(i + 2)] = x[3]
            ws1['E' + str(i + 2)] = x[4]
            ws1['F' + str(i + 2)] = x[5]
            ws1['G' + str(i + 2)] = x[6]
            ws1['H' + str(i + 2)] = x[7]
        cur.close()
        con.close()
    except sqlite3.Error as e:
        print('sqlite3 error occur In extract: %s', e.args[0])
    wb.save(filename)


def savesql(content, filename):
    con = sqlite3.connect(filename)
    cur = con.cursor()
    try:
        cur.execute('''create table if not exists Results
        (id             INTEGER     PRIMARY KEY     AUTOINCREMENT,
         nodeId         CHAR(2)     NOT NULL,
         rssi           INT         NOT NULL,
         seqNo          INT         NOT NULL,
         gettime        TEXT        NOT NULL,
         volt           FLOAT       NOT NULL,
         hopCount       INT         NOT NULL,
         timeDifference FLOAT       NOT NULL
         );
        ''')
    except sqlite3.Error as e:
        print('sqlite3 error occur In create: %s', e.args[0])

    temp = []
    for x in content:
        if x == '':
            break
        field = x.split(" ")
        if int(field[6], 16) > 128:
            RSSI_Val = -45 + int(field[6], 16) - 256
        else:
            RSSI_Val = -45 + int(field[6], 16)
        Voltage = int("".join(field[7:9]), 16) * 3.6 / 8192
        # timedifference需要请教老师
        if hex(int("".join(field[-1:9:-1]), 32)) > hex(int('7FFFFFFF', 32)):
            timedifference = float(int('FFFFFFFF', 32)) - float(int("".join(field[-1:9:-1]), 32))
        else:
            timedifference = float(int("".join(field[-1:9:-1]), 32))

        nodeId = "".join(field[2:4])
        rssi = float(RSSI_Val)
        seqNo = int("".join(field[4:6]), 16)
        gettime = str(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        hopCount = int("".join(field[9]), 16)
        temp.append((nodeId, rssi, seqNo, gettime, Voltage, hopCount, timedifference))
    try:
        for t in temp:
            cur.execute('''insert into Results(nodeId,rssi,seqNo,gettime,volt,hopCount,timeDifference) 
                                values(?,?,?,?,?,?,?)''', t)
    except sqlite3.Error as e:
        print("An sqlite3 error occurred insert: %s", e.args[0])
    finally:
        con.commit()
        cur.close()
        con.close()


def Statistic(filename):
    con = sqlite3.connect(filename)
    cur = con.cursor()
    try:
        cur.execute('''create table if not exists Statistic
        (ID             INTEGER     PRIMARY KEY     AUTOINCREMENT,
         nodeId         CHAR(2)     NOT NULL,
         lossPkgCount   INT         NOT NULL,
         lossPkgRate    FLOAT       NOT NULL,
         avgRSSI        FLOAT       NOT NULL,
         voltDiff       FLOAT       NOT NULL,
         avgHopCount    INT         NOT NULL,
         TimeDiff       FLOAT       NOT NULL
         );
        ''')
    except sqlite3.Error as e:
        print('sqlite3 error occur In create: %s', e.args[0])

    rows = cur.execute("SELECT * FROM Results ORDER BY nodeId ASC,seqNo ASC;").fetchall()
    row_seq = []
    row_seq.append(rows[0][1])
    for i in range(1, len(rows)):
        if rows[i][1] != rows[i - 1][1]:
            row_seq.append(rows[i][1])

    for i, x in enumerate(row_seq):
        x = str(x)
        max_volt = cur.execute("SELECT max(volt) FROM Results WHERE nodeId = '" + x + "';").fetchall()[0][0]
        min_volt = cur.execute("SELECT min(volt) FROM Results WHERE nodeId = '" + x + "';").fetchall()[0][0]
        voltDiff = max_volt - min_volt
        avgRSSI = cur.execute("SELECT avg(rssi) FROM Results WHERE nodeId = '" + x + "';").fetchall()[0][0]
        avgHopCount = cur.execute("SELECT avg(hopCount) FROM Results WHERE nodeId = '" + x + "';").fetchall()[0][0]
        TimeDiff = cur.execute("SELECT avg(timeDifference) FROM Results WHERE nodeId = '" + x + "';").fetchall()[0][0]
        recvCount = cur.execute("SELECT count(*) FROM Results WHERE nodeId = '" + x + "';").fetchall()[0][0]
        lossPkgCount = 0
        seq_data = cur.execute("SELECT seqNo FROM Results WHERE nodeId = '" + x + "' ORDER BY seqNo ASC;").fetchall()
        for j in range(1, len(seq_data)):
            if seq_data[j - 1][0] + 1 != seq_data[j][0]:
                lossPkgCount += seq_data[j][0] - seq_data[j - 1][0] - 1

        row = [0 for _ in range(7)]
        row[0] = x  # nodeId
        row[1] = lossPkgCount  # lossPkgCount
        row[2] = float(lossPkgCount) / (recvCount + lossPkgCount)  # lossPkgRate
        row[3] = avgRSSI
        row[4] = voltDiff
        row[5] = avgHopCount
        row[6] = TimeDiff
        cur.execute(
            '''insert into Statistic(nodeId,lossPkgCount,lossPkgRate,avgRSSI,voltDiff,avgHopCount,TimeDiff) values(?,?,?,?,?,?,?)''',
            tuple(row))
    con.commit()
    cur.close()
    con.close()
    print("statistic complete")


def main():
    print('Welcome to use protId sscom convert to xlsx program, coded by Xu Ming')
    sys.argv.append('')
    assert len(sys.argv) >= 3, "can input only one parameter 'filename.txt'"
    # strvar = ""
    if len(sys.argv) > 3:
        analyze(sys.argv[1], sys.argv[2])
        # strvar += read(sys.argv[1],sys.argv[2])
    elif len(sys.argv) == 3:
        analyze(sys.argv[1])
        # strvar += read(sys.argv[1])


if __name__ == '__main__':
    main()
